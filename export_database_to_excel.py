import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "database.json")
OUTPUT_PATH = os.path.join(BASE_DIR, "data", "AeroSense_Database.xlsx")
PROJECT_ROOT_OUTPUT = os.path.join(BASE_DIR, "AeroSense_Database.xlsx")

# Color Palette
PRIMARY_COLOR = "1E3A8A"      # Deep Navy
ACCENT_COLOR = "0D9488"       # Teal
HEADER_FONT_COLOR = "FFFFFF"  # White
ALT_ROW_FILL = "F8FAFC"       # Light Slate
BORDER_COLOR = "CBD5E1"       # Light gray border
TITLE_FILL = "0F172A"         # Dark slate

header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

accent_fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
accent_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)

title_fill = PatternFill(start_color=TITLE_FILL, end_color=TITLE_FILL, fill_type="solid")
title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

alt_fill = PatternFill(start_color=ALT_ROW_FILL, end_color=ALT_ROW_FILL, fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border_side = Side(style="thin", color=BORDER_COLOR)
cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

data_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")

def load_db():
    if not os.path.exists(DB_PATH):
        return {
            "users": [],
            "familyRequests": [],
            "notifications": [],
            "smsAlerts": [],
            "kidsProfiles": [],
            "locations": []
        }
    with open(DB_PATH, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except Exception:
            return {}

def auto_fit_columns(ws, min_width=12, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # split by newlines if any
            lines = val_str.split('\n')
            for l in lines:
                if len(l) > max_len:
                    max_len = len(l)
        adapted_width = max(min_width, min(max_len + 4, max_width))
        ws.column_dimensions[col_letter].width = adapted_width

def style_table_headers(ws, start_row, headers, fill=header_fill, font=header_font):
    ws.row_dimensions[start_row].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = header_align
        cell.border = cell_border

def create_excel():
    db = load_db()
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Sheet 1: Database Overview & Stats
    # -------------------------------------------------------------
    ws_overview = wb.active
    ws_overview.title = "Overview & Summary"
    ws_overview.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_overview.merge_cells("A1:E2")
    ws_overview["A1"] = "🌿 AeroSense System Database - Central Data Catalog"
    ws_overview["A1"].font = title_font
    ws_overview["A1"].fill = title_fill
    ws_overview["A1"].alignment = center_align
    
    ws_overview["A4"] = f"Export Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_overview["A4"].font = Font(name="Calibri", size=10, italic=True, color="475569")
    ws_overview["A5"] = "Application: AeroSense Real-Time Air Quality & Infection Susceptibility Tracker"
    ws_overview["A5"].font = Font(name="Calibri", size=10, bold=True, color="1E293B")
    
    overview_headers = ["Table Name", "Description", "Entity Key", "Total Records", "Status"]
    style_table_headers(ws_overview, 7, overview_headers, fill=accent_fill)
    
    overview_rows = [
        ["Users", "Registered system users, health profiles, credentials & emergency contacts", "User ID / Email", len(db.get("users", [])), "Active"],
        ["Kids_Profiles", "School routes, commute timings, and vulnerability profiles for children", "Kid ID / Parent ID", len(db.get("kidsProfiles", [])), "Active"],
        ["Family_Connections", "Family network requests and live tracking permissions", "Request ID", len(db.get("familyRequests", [])), "Active"],
        ["Notifications", "AQI threshold breach alerts, high-risk recommendations, and messages", "Notification ID", len(db.get("notifications", [])), "Active"],
        ["SMS_Alerts", "Emergency SMS dispatches sent to user & family emergency contacts", "Alert ID", len(db.get("smsAlerts", [])), "Active"],
        ["Live_Locations", "Real-time GPS telemetry and geolocation timestamps for risk tracking", "User ID", len(db.get("locations", [])), "Active"],
        ["AQI_Parameters_Ref", "Standardized air quality thresholds, pollutant weights & risk indices", "Parameter Code", 6, "Reference Model"],
        ["Data_Dictionary", "Complete schema data types, nullability, and field definitions", "Field Name", 32, "Schema Spec"]
    ]
    
    for row_idx, r_data in enumerate(overview_rows, 8):
        ws_overview.row_dimensions[row_idx].height = 22
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r_data, 1):
            c = ws_overview.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill
            c.font = bold_font if col_idx in (1, 4) else data_font
            c.alignment = center_align if col_idx in (3, 4, 5) else left_align
            c.border = cell_border
            
    auto_fit_columns(ws_overview)

    # -------------------------------------------------------------
    # Sheet 2: Users
    # -------------------------------------------------------------
    ws_users = wb.create_sheet(title="Users")
    ws_users.views.sheetView[0].showGridLines = True
    
    user_headers = [
        "User ID", "Full Name", "Email Address", "Role", "Phone", "Age", "Gender",
        "Blood Group", "Health Issues / Vulnerabilities", "Emergency Contact Name",
        "Emergency Contact Phone", "Notifications Enabled", "Location Sharing", "Theme"
    ]
    style_table_headers(ws_users, 1, user_headers)
    
    users_list = db.get("users", [])
    for row_idx, u in enumerate(users_list, 2):
        ws_users.row_dimensions[row_idx].height = 22
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        health_issues = u.get("healthIssues", [])
        if isinstance(health_issues, list):
            health_str = ", ".join(health_issues) if health_issues else "None specified"
        else:
            health_str = str(health_issues)
            
        settings = u.get("settings", {})
        
        row_vals = [
            u.get("id", ""),
            u.get("name", ""),
            u.get("email", ""),
            u.get("role", "user"),
            u.get("phone", ""),
            u.get("age", ""),
            u.get("gender", ""),
            u.get("bloodGroup", ""),
            health_str,
            u.get("emergencyContactName", ""),
            u.get("emergencyContactPhone", ""),
            "TRUE" if settings.get("notifications", True) else "FALSE",
            "TRUE" if settings.get("locationSharing", False) else "FALSE",
            settings.get("theme", "blue-white")
        ]
        
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_users.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill
            c.font = data_font
            c.alignment = center_align if col_idx in (1, 4, 6, 7, 8, 12, 13, 14) else left_align
            c.border = cell_border
            
    auto_fit_columns(ws_users)

    # -------------------------------------------------------------
    # Sheet 3: Kids_Profiles
    # -------------------------------------------------------------
    ws_kids = wb.create_sheet(title="Kids_Profiles")
    ws_kids.views.sheetView[0].showGridLines = True
    
    kids_headers = [
        "Kid ID", "Parent User ID", "Child Name", "Age", "School Name",
        "School Latitude", "School Longitude", "Commute Mode", "Departure Time",
        "Return Time", "Asthma / Respiratory Severity", "SMS Alerts Enabled", "Created Timestamp"
    ]
    style_table_headers(ws_kids, 1, kids_headers)
    
    kids_list = db.get("kidsProfiles", [])
    if not kids_list:
        # Provide clean placeholder row explaining table schema
        ws_kids.cell(row=2, column=1, value="No records stored currently - schema active for kid commute and vulnerability monitoring.").font = Font(italic=True, color="64748B")
    else:
        for row_idx, k in enumerate(kids_list, 2):
            ws_kids.row_dimensions[row_idx].height = 22
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            row_vals = [
                k.get("id", ""),
                k.get("parentId", ""),
                k.get("name", ""),
                k.get("age", ""),
                k.get("schoolName", ""),
                k.get("schoolLat", ""),
                k.get("schoolLon", ""),
                k.get("commuteMode", ""),
                k.get("departureTime", ""),
                k.get("returnTime", ""),
                k.get("asthmaLevel", "None"),
                "TRUE" if k.get("alertsEnabled", True) else "FALSE",
                k.get("createdAt", "")
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws_kids.cell(row=row_idx, column=col_idx, value=val)
                c.fill = fill
                c.font = data_font
                c.alignment = center_align if col_idx in (1, 2, 4, 6, 7, 8, 9, 10, 12) else left_align
                c.border = cell_border
                
    auto_fit_columns(ws_kids)

    # -------------------------------------------------------------
    # Sheet 4: Family_Connections
    # -------------------------------------------------------------
    ws_family = wb.create_sheet(title="Family_Connections")
    ws_family.views.sheetView[0].showGridLines = True
    
    family_headers = ["Request ID", "Requester User ID", "Requester Name", "Target User ID / Email / Phone", "Relationship Status", "Created Timestamp"]
    style_table_headers(ws_family, 1, family_headers)
    
    fam_list = db.get("familyRequests", [])
    if not fam_list:
        ws_family.cell(row=2, column=1, value="No family connections stored currently - schema ready for family safety network.").font = Font(italic=True, color="64748B")
    else:
        for row_idx, fam in enumerate(fam_list, 2):
            ws_family.row_dimensions[row_idx].height = 22
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            row_vals = [
                fam.get("id", ""),
                fam.get("from", ""),
                fam.get("fromName", fam.get("requesterName", "")),
                fam.get("to", fam.get("targetEmail", "")),
                fam.get("status", "pending"),
                fam.get("createdAt", "")
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws_family.cell(row=row_idx, column=col_idx, value=val)
                c.fill = fill
                c.font = data_font
                c.alignment = center_align if col_idx in (1, 2, 4, 5, 6) else left_align
                c.border = cell_border
                
    auto_fit_columns(ws_family)

    # -------------------------------------------------------------
    # Sheet 5: Notifications
    # -------------------------------------------------------------
    ws_notif = wb.create_sheet(title="Notifications")
    ws_notif.views.sheetView[0].showGridLines = True
    
    notif_headers = ["Notification ID", "Target User ID", "Category / Type", "Title", "Message Content", "Read Status", "Timestamp"]
    style_table_headers(ws_notif, 1, notif_headers)
    
    notif_list = db.get("notifications", [])
    if notif_list:
        for row_idx, n in enumerate(notif_list, 2):
            ws_notif.row_dimensions[row_idx].height = 22
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            row_vals = [
                n.get("id", ""),
                n.get("userId", ""),
                n.get("type", "warning"),
                n.get("title", ""),
                n.get("message", ""),
                "READ" if n.get("read", False) else "UNREAD",
                n.get("createdAt", n.get("timestamp", ""))
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws_notif.cell(row=row_idx, column=col_idx, value=val)
                c.fill = fill
                c.font = data_font
                c.alignment = center_align if col_idx in (1, 2, 3, 6, 7) else left_align
                c.border = cell_border
    else:
        ws_notif.cell(row=2, column=1, value="No notifications stored currently - dynamic push notifications trigger on AQI index > 150.").font = Font(italic=True, color="64748B")
        
    auto_fit_columns(ws_notif)

    # -------------------------------------------------------------
    # Sheet 6: SMS_Alerts
    # -------------------------------------------------------------
    ws_sms = wb.create_sheet(title="SMS_Alerts")
    ws_sms.views.sheetView[0].showGridLines = True
    
    sms_headers = ["Alert ID", "Sender User ID", "Destination Phone", "Alert Type", "SMS Message Body", "Delivery Status", "Timestamp"]
    style_table_headers(ws_sms, 1, sms_headers)
    
    sms_list = db.get("smsAlerts", [])
    if sms_list:
        for row_idx, s in enumerate(sms_list, 2):
            ws_sms.row_dimensions[row_idx].height = 22
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            row_vals = [
                s.get("id", ""),
                s.get("fromUserId", s.get("userId", "")),
                s.get("toPhone", s.get("phone", "")),
                s.get("type", "AQI_RISK_WARNING"),
                s.get("message", ""),
                s.get("status", "sent"),
                s.get("timestamp", s.get("createdAt", ""))
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws_sms.cell(row=row_idx, column=col_idx, value=val)
                c.fill = fill
                c.font = data_font
                c.alignment = center_align if col_idx in (1, 2, 3, 4, 6, 7) else left_align
                c.border = cell_border
    else:
        ws_sms.cell(row=2, column=1, value="No SMS alerts dispatched yet - real-time simulated & gateway SMS alerts are logged here.").font = Font(italic=True, color="64748B")
        
    auto_fit_columns(ws_sms)

    # -------------------------------------------------------------
    # Sheet 7: Live_Locations
    # -------------------------------------------------------------
    ws_loc = wb.create_sheet(title="Live_Locations")
    ws_loc.views.sheetView[0].showGridLines = True
    
    loc_headers = ["User ID", "User Display Name", "Latitude", "Longitude", "Accuracy (m)", "Speed (km/h)", "Heading (°)", "Last Synced Time"]
    style_table_headers(ws_loc, 1, loc_headers)
    
    loc_list = db.get("locations", [])
    if loc_list:
        for row_idx, loc in enumerate(loc_list, 2):
            ws_loc.row_dimensions[row_idx].height = 22
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            row_vals = [
                loc.get("userId", ""),
                loc.get("userName", loc.get("name", "")),
                loc.get("latitude", loc.get("lat", "")),
                loc.get("longitude", loc.get("lon", "")),
                loc.get("accuracy", 10),
                loc.get("speed", 0),
                loc.get("heading", 0),
                loc.get("timestamp", loc.get("updatedAt", ""))
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws_loc.cell(row=row_idx, column=col_idx, value=val)
                c.fill = fill
                c.font = data_font
                c.alignment = center_align if col_idx in (1, 3, 4, 5, 6, 7, 8) else left_align
                c.border = cell_border
    else:
        ws_loc.cell(row=2, column=1, value="No GPS locations registered yet - live device coordinates are populated on user location sharing.").font = Font(italic=True, color="64748B")
        
    auto_fit_columns(ws_loc)

    # -------------------------------------------------------------
    # Sheet 8: AQI_Parameters_Ref (Environmental Risk Scale)
    # -------------------------------------------------------------
    ws_aqi = wb.create_sheet(title="AQI_Parameters_Ref")
    ws_aqi.views.sheetView[0].showGridLines = True
    
    aqi_headers = [
        "AQI Range (CPCB / EPA)", "Risk Level", "PM2.5 (µg/m³)", "PM10 (µg/m³)",
        "Infection Susceptibility Multiplier", "Cardiopulmonary Risk Impact", "System Advisory / Recommended Action"
    ]
    style_table_headers(ws_aqi, 1, aqi_headers, fill=header_fill)
    
    aqi_data = [
        ["0 - 50", "Good", "0 - 30", "0 - 50", "1.0x (Baseline)", "Minimal risk", "Ideal air quality; suitable for all outdoor activities and exercise."],
        ["51 - 100", "Satisfactory / Moderate", "31 - 60", "51 - 100", "1.15x - 1.25x", "Minor breathing discomfort to sensitive individuals", "Acceptable air quality; sensitive individuals should monitor outdoor exposure."],
        ["101 - 200", "Moderate / Unhealthy Sensitive", "61 - 90", "101 - 250", "1.40x - 1.65x", "Breathing discomfort to people with lung disease, asthma, and children", "Kids & elderly should reduce heavy outdoor exertion; keep rescue inhalers ready."],
        ["201 - 300", "Poor / Unhealthy", "91 - 120", "251 - 350", "1.80x - 2.20x", "Breathing discomfort to most people on prolonged exposure", "Wear N95 masks outdoors; avoid morning runs; activate indoor air filtration."],
        ["301 - 400", "Very Poor", "121 - 250", "351 - 430", "2.50x - 3.20x", "Respiratory illness to the people on prolonged exposure", "Avoid all outdoor activities; close windows; enable high-efficiency HEPA purifiers."],
        ["401 - 500+", "Severe / Hazardous", "> 250", "> 430", "3.50x - 5.00x", "Affects healthy people and severely impacts those with existing diseases", "Emergency public health warning; stay strictly indoors; trigger automatic SMS alerts."]
    ]
    
    for row_idx, a_row in enumerate(aqi_data, 2):
        ws_aqi.row_dimensions[row_idx].height = 26
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(a_row, 1):
            c = ws_aqi.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill
            c.font = bold_font if col_idx in (1, 2, 5) else data_font
            c.alignment = center_align if col_idx in (1, 2, 3, 4, 5) else left_align
            c.border = cell_border
            
    auto_fit_columns(ws_aqi)

    # -------------------------------------------------------------
    # Sheet 9: Data_Dictionary
    # -------------------------------------------------------------
    ws_dict = wb.create_sheet(title="Data_Dictionary")
    ws_dict.views.sheetView[0].showGridLines = True
    
    dict_headers = ["Table Name", "Field Name", "Data Type", "Constraints", "Description"]
    style_table_headers(ws_dict, 1, dict_headers, fill=accent_fill)
    
    dict_data = [
        # Users
        ["Users", "id", "UUID (String)", "Primary Key, Unique, Not Null", "Globally unique identifier for the user account"],
        ["Users", "name", "VARCHAR(100)", "Not Null", "Full name of the registered user"],
        ["Users", "email", "VARCHAR(150)", "Unique, Not Null", "Email address used for authentication and communications"],
        ["Users", "passwordHash", "CHAR(60)", "Bcrypt Salted Hash", "Securely hashed password string"],
        ["Users", "role", "VARCHAR(20)", "Default: 'user' ('user' | 'expert' | 'admin')", "Access control level for application authorization"],
        ["Users", "phone", "VARCHAR(20)", "Optional", "Primary phone number for SMS emergency broadcast alerts"],
        ["Users", "age", "INTEGER", "Optional, Range: 1-120", "User age used for age-adjusted infection susceptibility modeling"],
        ["Users", "gender", "VARCHAR(20)", "Optional", "Gender identity of the user"],
        ["Users", "bloodGroup", "VARCHAR(10)", "Optional", "Blood type (e.g., A+, O+, B-, AB+) for medical emergency profile"],
        ["Users", "healthIssues", "ARRAY[String]", "Default: []", "Pre-existing health conditions (Asthma, COPD, Cardiac, Allergies)"],
        ["Users", "emergencyContactName", "VARCHAR(100)", "Optional", "Name of family member or emergency contact"],
        ["Users", "emergencyContactPhone", "VARCHAR(20)", "Optional", "Phone number for instantaneous SMS dispatch on hazardous AQI"],
        ["Users", "settings.notifications", "BOOLEAN", "Default: true", "Flag to enable in-app and push notification delivery"],
        ["Users", "settings.locationSharing", "BOOLEAN", "Default: false", "Opt-in consent for real-time family location tracking"],
        ["Users", "settings.theme", "VARCHAR(30)", "Default: 'blue-white'", "User UI theme preference"],

        # Kids Profiles
        ["Kids_Profiles", "id", "UUID (String)", "Primary Key, Unique", "Unique child profile record identifier"],
        ["Kids_Profiles", "parentId", "UUID (String)", "Foreign Key -> Users.id", "User ID of the parent or guardian"],
        ["Kids_Profiles", "name", "VARCHAR(100)", "Not Null", "Child's name"],
        ["Kids_Profiles", "age", "INTEGER", "Optional", "Child's age for paediatric respiratory vulnerability scoring"],
        ["Kids_Profiles", "schoolName", "VARCHAR(150)", "Not Null", "Name of school or learning institution"],
        ["Kids_Profiles", "schoolLat", "DECIMAL(10,6)", "Not Null", "Geographical latitude of the school location"],
        ["Kids_Profiles", "schoolLon", "DECIMAL(10,6)", "Not Null", "Geographical longitude of the school location"],
        ["Kids_Profiles", "commuteMode", "VARCHAR(30)", "e.g. Bus, Walking, Cycling, Car", "Method of daily commute for exposure modeling"],
        ["Kids_Profiles", "departureTime", "TIME (HH:MM)", "e.g. 07:30", "Daily morning school departure time"],
        ["Kids_Profiles", "returnTime", "TIME (HH:MM)", "e.g. 15:30", "Daily afternoon school return time"],
        ["Kids_Profiles", "asthmaLevel", "VARCHAR(30)", "'None' | 'Mild' | 'Moderate' | 'Severe'", "Severity of pediatric respiratory or asthmatic conditions"],
        ["Kids_Profiles", "alertsEnabled", "BOOLEAN", "Default: true", "Toggle for automatic school zone air quality alerts"],

        # Family Connections
        ["Family_Connections", "id", "UUID (String)", "Primary Key, Unique", "Unique connection request ID"],
        ["Family_Connections", "from", "UUID (String)", "Foreign Key -> Users.id", "Requester user ID initiating family network invite"],
        ["Family_Connections", "to", "UUID / String", "Target ID / Email", "Recipient user identifier or invited email"],
        ["Family_Connections", "status", "VARCHAR(20)", "'pending' | 'accepted' | 'rejected'", "Current approval status of family link"],
        ["Family_Connections", "createdAt", "TIMESTAMP", "ISO-8601 UTC", "Timestamp when the family connection was initiated"],

        # Notifications
        ["Notifications", "id", "UUID (String)", "Primary Key", "Unique alert identifier"],
        ["Notifications", "userId", "UUID (String)", "Foreign Key -> Users.id", "Target user ID"],
        ["Notifications", "type", "VARCHAR(30)", "'info' | 'warning' | 'danger' | 'emergency'", "Severity category of atmospheric risk"],
        ["Notifications", "title", "VARCHAR(150)", "Not Null", "Summary header of warning"],
        ["Notifications", "message", "TEXT", "Not Null", "Detailed advisory and health recommendations"],
        ["Notifications", "read", "BOOLEAN", "Default: false", "Indicator if user has viewed the notification"],
        ["Notifications", "createdAt", "TIMESTAMP", "ISO-8601 UTC", "Timestamp of notification dispatch"],

        # SMS Alerts
        ["SMS_Alerts", "id", "UUID (String)", "Primary Key", "Unique SMS dispatch log ID"],
        ["SMS_Alerts", "fromUserId", "UUID (String)", "Foreign Key -> Users.id", "User triggering or associated with alert"],
        ["SMS_Alerts", "toPhone", "VARCHAR(20)", "Not Null", "Recipient phone number receiving alert message"],
        ["SMS_Alerts", "message", "TEXT", "Not Null", "Text body sent via SMS gateway"],
        ["SMS_Alerts", "status", "VARCHAR(20)", "'sent' | 'delivered' | 'failed'", "Dispatch delivery status"],
        ["SMS_Alerts", "timestamp", "TIMESTAMP", "ISO-8601 UTC", "Timestamp when message was logged"],

        # Live Locations
        ["Live_Locations", "userId", "UUID (String)", "Primary Key, Foreign Key -> Users.id", "User ID associated with coordinates"],
        ["Live_Locations", "latitude", "DECIMAL(10,6)", "Not Null, Range: -90 to 90", "Real-time latitude coordinate"],
        ["Live_Locations", "longitude", "DECIMAL(10,6)", "Not Null, Range: -180 to 180", "Real-time longitude coordinate"],
        ["Live_Locations", "accuracy", "DECIMAL(6,2)", "Optional (Meters)", "GPS fix horizontal accuracy radius"],
        ["Live_Locations", "speed", "DECIMAL(6,2)", "Optional (km/h)", "Current travel speed"],
        ["Live_Locations", "timestamp", "TIMESTAMP", "ISO-8601 UTC", "Last updated coordinate timestamp"]
    ]
    
    for row_idx, d_row in enumerate(dict_data, 2):
        ws_dict.row_dimensions[row_idx].height = 20
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(d_row, 1):
            c = ws_dict.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill
            c.font = bold_font if col_idx in (1, 2) else data_font
            c.alignment = center_align if col_idx in (1, 3, 4) else left_align
            c.border = cell_border
            
    auto_fit_columns(ws_dict)

    # Save to both data/ folder and root directory for easy access
    wb.save(OUTPUT_PATH)
    wb.save(PROJECT_ROOT_OUTPUT)
    print(f"[SUCCESS] Excel database exported successfully to:\n1. {OUTPUT_PATH}\n2. {PROJECT_ROOT_OUTPUT}")

if __name__ == "__main__":
    create_excel()
